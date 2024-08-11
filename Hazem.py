import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from tkinter import ttk
from datetime import datetime
from collections import Counter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import requests
import zipfile
import subprocess
import sys
import shutil

class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("برنامج الورشه تم التطوير بواسطة حازم ايمن محمود")
        
        # تعيين أبعاد الشاشة ولون الخلفية مع مؤثرات بصرية
        self.root.geometry("500x600")
        self.root.configure(bg='#e6e6fa')
        
        # إضافة تدرج لون كخلفية للشاشة
        self.canvas = tk.Canvas(self.root, width=500, height=600)
        self.canvas.pack(fill="both", expand=True)
        
        self.gradient_background(self.canvas, '#e6e6fa', '#b0c4de')  # مؤثر التدرج
        
        self.file_path = ""
        
        # زر لاختيار ملف Excel
        self.select_button = tk.Button(self.canvas, text="اختر ملف Excel", command=self.select_file, bg='#b0c4de')
        self.select_button.pack(pady=10)
        
        # مربع نص لإدخال الكود
        self.code_label = tk.Label(self.canvas, text="أدخل الكود:", bg='#e6e6fa')
        self.code_label.pack()
        self.code_entry = tk.Entry(self.canvas)
        self.code_entry.pack(pady=5)
        self.code_entry.bind("<Return>", self.add_data)  # إضافة البيانات عند الضغط على Enter
        
        # قائمة منسدلة لاختيار اسم الفني
        self.technician_label = tk.Label(self.canvas, text="اختر اسم الفني:", bg='#e6e6fa')
        self.technician_label.pack(pady=1)
        self.technician_combo = ttk.Combobox(self.canvas, values=[
            "محمد عادل", "محمد الزغبى", "حسنى ضاحى", 
            "حسام محمد ابراهيم", "احمد عشرى", "محمد سمير مصطفى",
            "احمد عزازى", "كريم اشرف", "محمد حسن عامر", 
            "عبد الرحمن هلال", "عبد الرحمن"
        ])
        self.technician_combo.pack(pady=5)
        self.technician_combo.bind("<<ComboboxSelected>>", self.show_technician_details)  # عرض التفاصيل عند اختيار الفني
        
        # Checkboxes
        self.checkbox1_var = tk.BooleanVar()
        self.checkbox2_var = tk.BooleanVar()
        self.checkbox3_var = tk.BooleanVar()
        
        self.checkbox1 = tk.Checkbutton(self.canvas, text="داخل الضمان", variable=self.checkbox1_var, bg='#e6e6fa')
        self.checkbox1.pack(pady=2)
        self.checkbox2 = tk.Checkbutton(self.canvas, text="خارج الضمان ", variable=self.checkbox2_var, bg='#e6e6fa')
        self.checkbox2.pack(pady=2)
        self.checkbox3 = tk.Checkbutton(self.canvas, text="مرتجع", variable=self.checkbox3_var, bg='#e6e6fa')
        self.checkbox3.pack(pady=2)
        
        # حقل إدخال التاريخ
        self.date_label = tk.Label(self.canvas, text="أدخل التاريخ (YYYY-MM-DD):", bg='#e6e6fa')
        self.date_label.pack(pady=5)
        self.date_entry = tk.Entry(self.canvas)
        self.date_entry.pack(pady=5)
        self.date_entry.insert(0, datetime.today().strftime('%Y-%m-%d'))  # تعبئة التاريخ الحالي تلقائيًا
        
        # زر لإضافة البيانات
        self.add_button = tk.Button(self.canvas, text="إضافة البيانات", command=self.add_data, bg='#b0c4de')
        self.add_button.pack(pady=10)
        
        # زر لعرض شاشة البيانات الاحترافية
        self.view_button = tk.Button(self.canvas, text="عرض البيانات", command=self.view_data, bg='#b0c4de')
        self.view_button.pack(pady=10)
        
        # زر لعرض شاشة التقارير الاحترافية
        self.report_button = tk.Button(self.canvas, text="عرض التقارير", command=self.show_reports, bg='#b0c4de')
        self.report_button.pack(pady=10)

        # زر لعرض شاشة التعديل والحذف
        self.edit_delete_button = tk.Button(self.canvas, text="تعديل/حذف البيانات", command=self.show_edit_delete, bg='#b0c4de')
        self.edit_delete_button.pack(pady=10)

        # زر لتحديث البرنامج

        self.update_button = tk.Button(self.canvas, text="تحديث البرنامج", command=self.update_program, bg='#b0c4de')
        self.update_button.pack(pady=10)


    def gradient_background(self, canvas, color1, color2):
        """Create a gradient background."""
        for i in range(500):
            r = int(i * (int(color2[1:3], 16) - int(color1[1:3], 16)) / 500 + int(color1[1:3], 16))
            g = int(i * (int(color2[3:5], 16) - int(color1[3:5], 16)) / 500 + int(color1[3:5], 16))
            b = int(i * (int(color2[5:7], 16) - int(color1[5:7], 16)) / 500 + int(color1[5:7], 16))
            color = f'#{r:02x}{g:02x}{b:02x}'
            canvas.create_line(0, i, 500, i, fill=color)
    
    def add_scroll(self, window):
        """Add a scrollbar to a window."""
        container = tk.Frame(window)
        canvas = tk.Canvas(container)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        container.pack(fill="both", expand=True)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        return scrollable_frame
    
    def select_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            messagebox.showinfo("تم الاختيار", f"تم اختيار الملف: {self.file_path}")
    
    def add_data(self, event=None):
        if not self.file_path:
            messagebox.showerror("خطأ", "يرجى اختيار ملف Excel أولاً")
            return
        
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            # البحث عن أول صف فارغ
            row = 2
            while sheet[f"A{row}"].value is not None:
                row += 1
            
            # إضافة البيانات
            sheet[f"A{row}"] = row - 1
            sheet[f"B{row}"] = self.code_entry.get()
            sheet[f"E{row}"] = "True" if self.checkbox1_var.get() else ""
            sheet[f"F{row}"] = "True" if self.checkbox2_var.get() else ""
            sheet[f"G{row}"] = "True" if self.checkbox3_var.get() else ""
            sheet[f"H{row}"] = self.technician_combo.get()
            sheet[f"I{row}"] = self.date_entry.get()  # إضافة التاريخ
            
            wb.save(self.file_path)
            messagebox.showinfo("نجاح", "تمت إضافة البيانات بنجاح")
            
            # إعادة تعيين المدخلات
            self.code_entry.delete(0, tk.END)
            self.date_entry.delete(0, tk.END)
            self.date_entry.insert(0, datetime.today().strftime('%Y-%m-%d'))
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء حفظ البيانات: {e}")
    
    def view_data(self):
        if not self.file_path:
            messagebox.showerror("خطأ", "يرجى اختيار ملف Excel أولاً")
            return
        
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            technician_count = Counter()
            for row in range(2, sheet.max_row + 1):
                technician = sheet[f"H{row}"].value
                if technician:
                    technician_count[technician] += 1
            
            # إنشاء نافذة جديدة لعرض البيانات مع سكرول
            view_window = tk.Toplevel(self.root)
            view_window.title("عرض البيانات")
            view_window.geometry("400x300")
            
            scrollable_frame = self.add_scroll(view_window)
            
            # عرض البيانات
            for technician, count in technician_count.items():
                label = tk.Label(scrollable_frame, text=f"{technician}: {count} قطع")
                label.pack(pady=5)
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء عرض البيانات: {e}")
    
    def show_technician_details(self, event):
        if not self.file_path:
            messagebox.showerror("خطأ", "يرجى اختيار ملف Excel أولاً")
            return
        
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            technician_name = self.technician_combo.get()
            parts_count = 0
            parts_details = []
            for row in range(2, sheet.max_row + 1):
                if sheet[f"H{row}"].value == technician_name:
                    parts_count += 1
                    parts_details.append({
                        'name': sheet[f"B{row}"].value,
                        'price': sheet[f"C{row}"].value
                    })
            
            # إنشاء نافذة جديدة لعرض تفاصيل الفني
            details_window = tk.Toplevel(self.root)
            details_window.title("تفاصيل الفني")
            details_window.geometry("400x300")
            
            scrollable_frame = self.add_scroll(details_window)
            
            # عرض تفاصيل الفني
            count_label = tk.Label(scrollable_frame, text=f"عدد القطع: {parts_count}")
            count_label.pack(pady=5)
            
            for detail in parts_details:
                part_label = tk.Label(scrollable_frame, text=f"اسم القطعة: {detail['name']}, السعر: {detail['price']}")
                part_label.pack(pady=5)
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء عرض تفاصيل الفني: {e}")
    
    def show_edit_delete(self):
        if not self.file_path:
            messagebox.showerror("خطأ", "يرجى اختيار ملف Excel أولاً")
            return
        
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            # إنشاء نافذة جديدة لعرض شاشة التعديل والحذف
            edit_delete_window = tk.Toplevel(self.root)
            edit_delete_window.title("تعديل/حذف البيانات")
            edit_delete_window.geometry("600x400")
            
            scrollable_frame = self.add_scroll(edit_delete_window)
            
            # عرض بيانات التعديل والحذف
            for row in range(2, sheet.max_row + 1):
                part_name = sheet[f"B{row}"].value
                part_price = sheet[f"C{row}"].value
                
                # إضافة واجهة التعديل والحذف لكل قطعة
                label = tk.Label(scrollable_frame, text=f"اسم القطعة: {part_name}, السعر: {part_price}")
                label.pack(pady=5)
                delete_button = tk.Button(scrollable_frame, text="حذف", command=lambda r=row: self.delete_data(r))
                delete_button.pack(pady=5)
                edit_button = tk.Button(scrollable_frame, text="تعديل", command=lambda r=row: self.edit_data(r))
                edit_button.pack(pady=5)
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء عرض شاشة التعديل والحذف: {e}")
    
    def delete_data(self, row):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            # حذف البيانات من الصف المحدد
            sheet.delete_rows(row)
            
            wb.save(self.file_path)
            messagebox.showinfo("نجاح", "تم حذف البيانات بنجاح")
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء حذف البيانات: {e}")
    
    def edit_data(self, row):
        # واجهة التعديل (يمكنك تخصيصها حسب الحاجة)
        edit_window = tk.Toplevel(self.root)
        edit_window.title("تعديل البيانات")
        edit_window.geometry("400x300")
        
        # حقل لتعديل اسم القطعة
        name_label = tk.Label(edit_window, text="اسم القطعة:")
        name_label.pack(pady=5)
        name_entry = tk.Entry(edit_window)
        name_entry.pack(pady=5)
        
        # حقل لتعديل سعر القطعة
        price_label = tk.Label(edit_window, text="سعر القطعة:")
        price_label.pack(pady=5)
        price_entry = tk.Entry(edit_window)
        price_entry.pack(pady=5)
        
        save_button = tk.Button(edit_window, text="حفظ", command=lambda: self.save_edits(row, name_entry.get(), price_entry.get()))
        save_button.pack(pady=10)
    
    def save_edits(self, row, new_name, new_price):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            # تعديل البيانات في الصف المحدد
            sheet[f"B{row}"] = new_name
            sheet[f"C{row}"] = new_price
            
            wb.save(self.file_path)
            messagebox.showinfo("نجاح", "تم حفظ التعديلات بنجاح")
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء حفظ التعديلات: {e}")
    
    def show_reports(self):
        if not self.file_path:
            messagebox.showerror("خطأ", "يرجى اختيار ملف Excel أولاً")
            return
        
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            technician_count = Counter()
            for row in range(2, sheet.max_row + 1):
                technician = sheet[f"H{row}"].value
                if technician:
                    technician_count[technician] += 1
            
            # إنشاء نافذة جديدة للتقارير
            report_window = tk.Toplevel(self.root)
            report_window.title("التقارير")
            report_window.geometry("600x400")
            
            fig, ax = plt.subplots(figsize=(6, 4))
            ax.bar(technician_count.keys(), technician_count.values())
            ax.set_xlabel('اسم الفني')
            ax.set_ylabel('عدد القطع')
            ax.set_title('تقرير عدد القطع لكل فني')
            
            canvas = FigureCanvasTkAgg(fig, master=report_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء عرض التقارير: {e}")
    
    def update_program(self):
        """Download the latest version of the program and update it."""
        try:
            url = "https://github.com/hazem589799/Hazem5897999/archive/refs/heads/main.zip"
            local_zip_path = "update.zip"
            new_program_path = "Hazem5897999-main/Hazem.py"
            
            # تحميل الملف المضغوط
            response = requests.get(url)
            with open(local_zip_path, 'wb') as file:
                file.write(response.content)
            
            # فك ضغط الملف
            with zipfile.ZipFile(local_zip_path, 'r') as zip_ref:
                zip_ref.extractall()
            
            # تحويل الملف إلى exe
            subprocess.call([
                'pyinstaller', '--onefile', '--distpath', '.', '--workpath', '.', '--specpath', '.', new_program_path
            ])
            
            # اسم الملف التنفيذي الجديد
            exe_file = "Hazem.exe"
            
            # الحصول على مسار الملف الأصلي
            original_path = os.path.abspath(__file__)
            original_dir = os.path.dirname(original_path)
            
            # المسار الجديد للملف التنفيذي
            new_exe_path = os.path.join(original_dir, exe_file)
            
            # حذف النسخة القديمة
            if os.path.isfile(new_exe_path):
                os.remove(new_exe_path)
            
            # نقل النسخة الجديدة إلى المسار الأصلي
            shutil.move(os.path.join('.', exe_file), new_exe_path)
            
            # حذف الملف المضغوط والتحديث
            os.remove(local_zip_path)
            shutil.rmtree('Hazem5897999-main')
            
            # إعادة تشغيل البرنامج بالنسخة الجديدة
            subprocess.call([new_exe_path])
            self.root.destroy()
        
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء تحديث البرنامج: {e}")

        self.update_program()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()
